import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import linregress

# Chargement des données Excel
wb = openpyxl.load_workbook('PPG_data_combined.xlsx')
ws_sheet1 = wb['Sheet1']
ws_sheet2 = wb['Sheet2']

# Fonction pour extraire les données d'une feuille
def extract_data(sheet, columns, start_row=2):
    return [
        [sheet.cell(row=i, column=col).value for i in range(start_row, sheet.max_row + 1)]
        for col in columns
    ]

# Extraction des données
sheet1_columns = [2, 4, 6, 8, 10, 12, 14, 16]
sheet2_columns = [3, 5, 7, 9, 10, 11, 12, 13, 14]

sheet1_data = extract_data(ws_sheet1, sheet1_columns)
sheet2_data = extract_data(ws_sheet2, sheet2_columns)

# Association des données avec leurs labels
datasets = {
    'Ranked1': sheet1_data[0], 'Ranked2': sheet1_data[1], 'Ranked3': sheet1_data[2],
    'Ranked4': sheet1_data[3], 'Ranked5': sheet1_data[4], 'Ranked6': sheet1_data[5],
    'Ranked7': sheet1_data[6], 'Reference': sheet1_data[7],
    'Normal1': sheet2_data[0], 'Normal2': sheet2_data[1], 'Normal3': sheet2_data[2],
    'Normal4': sheet2_data[3], 'Ranked8': sheet2_data[4], 'Ranked9': sheet2_data[5],
    'Ranked10': sheet2_data[6], 'Ranked11': sheet2_data[7], 'Ranked12': sheet2_data[8],
}

# Calcul des moyennes
means = {label: np.mean([x for x in data if x is not None]) for label, data in datasets.items()}

# Fonction pour calculer la variabilité cumulée
def compute_cumulative_variability(data, mean):
    cumulative_variability = []
    current_variability = 0
    for value in data:
        if value is not None:
            current_variability += abs(value - mean)
            cumulative_variability.append(current_variability)
    return cumulative_variability

# Calcul des variabilités cumulées et des pentes
cumulative_variabilities = {
    label: compute_cumulative_variability(data, means[label]) for label, data in datasets.items()
}
slopes = {
    label: linregress(range(len(data)), data).slope if len(data) > 1 else None
    for label, data in cumulative_variabilities.items()
}

# Fenêtre 1 : Graphiques pour toutes les séries "Ranked"
ranked_labels = ['Ranked1', 'Ranked2', 'Ranked3', 'Ranked4', 'Ranked5', 
                 'Ranked6', 'Ranked7', 'Ranked8', 'Ranked9', 'Ranked10', 
                 'Ranked11', 'Ranked12']

fig1, axs1 = plt.subplots(len(ranked_labels), 1, figsize=(12, len(ranked_labels) * 2), sharex=True)
colors1 = plt.cm.tab10(np.linspace(0, 1, len(ranked_labels)))

for i, label in enumerate(ranked_labels):
    data = datasets[label]
    valid_indices = [j for j, x in enumerate(data) if x is not None]
    valid_data = [data[j] for j in valid_indices]
    mean_value = means[label]
    variability = sum(abs(x - mean_value) for x in valid_data)

    axs1[i].plot(valid_indices, valid_data, label=label, color=colors1[i])
    axs1[i].axhline(y=mean_value, color='gray', linestyle='--', label=f'Mean: {mean_value:.2f}')
    axs1[i].fill_between(valid_indices, valid_data, mean_value, color=colors1[i], alpha=0.3)
    axs1[i].set_title(f'{label} (Mean: {mean_value:.2f}, Variability: {variability:.2f})')
    axs1[i].legend()
    axs1[i].grid(True)

plt.tight_layout()

# Fenêtre 2 : Graphiques pour "Normal" et "Reference"
normal_and_reference_labels = ['Normal1', 'Normal2', 'Normal3', 'Normal4', 'Reference']

fig2, axs2 = plt.subplots(len(normal_and_reference_labels), 1, figsize=(12, len(normal_and_reference_labels) * 2), sharex=True)
colors2 = plt.cm.tab10(np.linspace(0, 1, len(normal_and_reference_labels)))

for i, label in enumerate(normal_and_reference_labels):
    data = datasets[label]
    valid_indices = [j for j, x in enumerate(data) if x is not None]
    valid_data = [data[j] for j in valid_indices]
    mean_value = means[label]
    variability = sum(abs(x - mean_value) for x in valid_data)

    axs2[i].plot(valid_indices, valid_data, label=label, color=colors2[i])
    axs2[i].axhline(y=mean_value, color='gray', linestyle='--', label=f'Mean: {mean_value:.2f}')
    axs2[i].fill_between(valid_indices, valid_data, mean_value, color=colors2[i], alpha=0.3)
    axs2[i].set_title(f'{label} (Mean: {mean_value:.2f}, Variability: {variability:.2f})')
    axs2[i].legend()
    axs2[i].grid(True)

plt.tight_layout()

# Fenêtre 3 : Variabilité cumulée pour tous les datasets
fig3, ax3 = plt.subplots(figsize=(12, 10))
for (label, variability), color in zip(cumulative_variabilities.items(), plt.cm.tab10(np.linspace(0, 1, len(datasets)))):
    ax3.plot(range(len(variability)), variability, label=f'{label} (Slope: {slopes[label]:.4f})', color=color)

ax3.set_title('Cumulative Variability with Slopes in Legend (All Datasets)')
ax3.set_xlabel('Time')
ax3.set_ylabel('Cumulative Variability')
ax3.legend()
ax3.grid(True)

plt.tight_layout()

# Affichage des graphiques
plt.show()
