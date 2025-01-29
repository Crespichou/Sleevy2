import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import linregress

# Chargement des données Excel
wb = openpyxl.load_workbook('PPG_data_combined.xlsx')
ws_sheet1 = wb['Sheet1']
ws_sheet2 = wb['Sheet2']

# Fonction pour extraire les données d'une feuille
def extract_data(sheet, columns, start_row=2):
    return [
        [sheet.cell(row=i, column=col).value for i in range(start_row, sheet.max_row + 1)]
        for col in columns
    ]

# Extraction des données
sheet1_columns = [2, 4, 6, 8, 10, 12, 14, 16]
sheet2_columns = [10, 11, 12, 13]

sheet1_data = extract_data(ws_sheet1, sheet1_columns)
sheet2_data = extract_data(ws_sheet2, sheet2_columns)

# Association des données avec leurs labels
datasets = {
    'Ranked1': sheet1_data[0], 'Ranked2': sheet1_data[1], 'Ranked3': sheet1_data[2],
    'Ranked4': sheet1_data[3], 'Ranked5': sheet1_data[4], 'Ranked6': sheet1_data[5],
    'Ranked7': sheet1_data[6], 'Reference': sheet1_data[7],
    'Ranked8': sheet2_data[0], 'Ranked9': sheet2_data[1],
    'Ranked10': sheet2_data[2], 'Ranked11': sheet2_data[3]
}

# Calcul de la moyenne fixe basée sur le jeu de donnée de référence (au repos)
reference_data = [x for x in datasets['Reference'] if x is not None]
reference_mean = np.mean(reference_data)

# Fonction pour calculer la variabilité cumulée en utilisant la moyenne fixe (aire entre courbe et moyenne)
def compute_cumulative_variability(data, fixed_mean):
    cumulative_variability = []
    current_variability = 0
    for value in data:
        if value is not None:
            current_variability += abs(value - fixed_mean)
            cumulative_variability.append(current_variability)
    return cumulative_variability

# Calcul des variabilités cumulées avec la moyenne de référence
cumulative_variabilities = {
    label: compute_cumulative_variability(data, reference_mean) for label, data in datasets.items()
}

# Calcul de la variabilité totale pour chaque série
variabilities = {
    label: sum(abs(x - reference_mean) for x in data if x is not None)
    for label, data in datasets.items()
}

# Définir une tolérance (exemple : ±10% de la variabilité moyenne)
tolerance = 0.1 * np.mean(list(variabilities.values()))

# Regrouper les séries par variabilité communes et créer des groupes
groups = {}
for label, variability in variabilities.items():
    added_to_group = False
    for group_variability in groups:
        if abs(variability - group_variability) <= tolerance:
            groups[group_variability].append(label)
            added_to_group = True
            break
    if not added_to_group:
        groups[variability] = [label]
        
# Fenêtre  : Graphiques pour toutes les séries "Ranked"
ranked_labels = ['Ranked1', 'Ranked2', 'Ranked3', 'Ranked4', 'Ranked5', 
                 'Ranked6', 'Ranked7', 'Ranked8', 'Ranked9', 'Ranked10', 
                 'Ranked11']

fig1, axs1 = plt.subplots(len(ranked_labels), 1, figsize=(12, len(ranked_labels) * 2), sharex=True)
colors1 = plt.cm.tab10(np.linspace(0, 1, len(ranked_labels)))

for i, label in enumerate(ranked_labels):
    data = datasets[label]
    valid_indices = [j for j, x in enumerate(data) if x is not None]
    valid_data = [data[j] for j in valid_indices]
    variability = sum(abs(x - reference_mean) for x in valid_data)

    axs1[i].plot(valid_indices, valid_data, label=label, color=colors1[i])
    axs1[i].axhline(y=reference_mean, color='gray', linestyle='--', label=f'Ref. Mean: {reference_mean:.2f}')
    axs1[i].fill_between(valid_indices, valid_data, reference_mean, color=colors1[i], alpha=0.3)
    axs1[i].set_title(f'{label} (Ref. Mean: {reference_mean:.2f}, Variability: {variability:.2f})')
    axs1[i].legend()
    axs1[i].grid(True)

plt.tight_layout()

# Afficher les groupes formés
for i, (group_variability, labels) in enumerate(groups.items(), 1):
    print(f"Groupe {i} (Variabilité ~ {group_variability:.2f}): {', '.join(labels)}")

# Fenêtre : Graphiques pour chaque groupe (évolution des variabilités cumulées)
fig1, axs1 = plt.subplots(len(groups), 1, figsize=(12, len(groups) * 4), sharex=True)
colors1 = plt.cm.tab10(np.linspace(0, 1, len(groups)))

# Variabilité cumulée de la référence
reference_cumulative_variability = cumulative_variabilities['Reference']

for i, (group_variability, labels) in enumerate(groups.items()):
    ax = axs1[i] if len(groups) > 1 else axs1  # Gérer le cas d'un seul groupe
    for label in labels:
        # Récupération de la variabilité cumulée pour chaque série du groupe
        group_cumulative_variability = cumulative_variabilities[label]
        ax.plot(
            range(len(group_cumulative_variability)),
            group_cumulative_variability,
            label=f"{label} (Var: {variabilities[label]:.2f})",
            color=colors1[i],
        )
    
    # Ajout de la variabilité cumulée de la référence
    ax.plot(
        range(len(reference_cumulative_variability)),
        reference_cumulative_variability,
        label="Référence (Cumulative)",
        color='black',
        linestyle='--'
    )
    
    ax.set_title(f"Groupe {i+1} (Variabilité ~ {group_variability:.2f})")
    ax.set_xlabel("Temps")
    ax.set_ylabel("Variabilité cumulée")
    ax.legend()
    ax.grid(True)

plt.tight_layout()

# Affichage des graphiques
plt.show()
