import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Fonction de chargement des données
def load_data(file_path, sheet_names, columns_per_sheet):
    """Charge les données d'un fichier Excel pour des colonnes spécifiques."""
    wb = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name, columns in zip(sheet_names, columns_per_sheet):
        sheet = wb[sheet_name]
        data[sheet_name] = extract_data(sheet, columns)
    return data

# Fonction d'extraction des données
def extract_data(sheet, columns, start_row=2):
    """Extrait les données d'une feuille pour des colonnes spécifiques."""
    return [
        [sheet.cell(row=i, column=col).value for i in range(start_row, sheet.max_row + 1)]
        for col in columns
    ]

# Calcul de la moyenne de référence
def calculate_reference_mean(reference_data):
    """Calcule la moyenne des données de référence en ignorant les valeurs None."""
    return np.mean([x for x in reference_data if x is not None])

# Calcul de la variabilité cumulée
def compute_cumulative_variability(data, fixed_mean):
    """Calcule la variabilité cumulée par rapport à une moyenne fixe."""
    cumulative_variability = []
    current_variability = 0
    for value in data:
        if value is not None:
            current_variability += abs(value - fixed_mean)
            cumulative_variability.append(current_variability)
    return cumulative_variability

# Calcul des variabilités
def calculate_variabilities(datasets, reference_mean):
    """Calcule la variabilité totale pour chaque série de données."""
    return {
        label: sum(abs(x - reference_mean) for x in data if x is not None)
        for label, data in datasets.items()
    }

# Regroupement des données par variabilité
def group_datasets_by_variability(variabilities, tolerance):
    """Regroupe les séries en fonction de leur variabilité avec une tolérance donnée."""
    groups = {}
    for label, variability in variabilities.items():
        added_to_group = False
        for group_variability in groups:
            if abs(variability - group_variability) <= tolerance:
                groups[group_variability].append(label)
                added_to_group = True
                break
        if not added_to_group:
            groups[variability] = [label]
    return groups

# Tracé des séries classées
def plot_ranked_series(datasets, reference_mean, ranked_labels):
    """Trace les séries classées (ranked) avec la variabilité par rapport à la référence."""
    fig, axs = plt.subplots(len(ranked_labels), 1, figsize=(12, len(ranked_labels) * 2), sharex=True)
    colors = plt.cm.tab10(np.linspace(0, 1, len(ranked_labels)))

    for i, label in enumerate(ranked_labels):
        data = datasets[label]
        valid_indices = [j for j, x in enumerate(data) if x is not None]
        valid_data = [data[j] for j in valid_indices]
        variability = sum(abs(x - reference_mean) for x in valid_data)

        axs[i].plot(valid_indices, valid_data, label=label, color=colors[i])
        axs[i].axhline(y=reference_mean, color='gray', linestyle='--', label=f'Ref. Mean: {reference_mean:.2f}')
        axs[i].fill_between(valid_indices, valid_data, reference_mean, color=colors[i], alpha=0.3)
        axs[i].set_title(f'{label} (Ref. Mean: {reference_mean:.2f}, Variability: {variability:.2f})')
        axs[i].legend()
        axs[i].grid(True)

    plt.tight_layout()
    # Commenter la ligne suivante pour éviter d'afficher le graphique
    #plt.show()

# Tracé des variabilités cumulées pour chaque groupe
def plot_grouped_cumulative_variability(groups, cumulative_variabilities, reference_cumulative_variability, variabilities):
    """Trace la variabilité cumulée pour chaque groupe."""
    fig, axs = plt.subplots(len(groups), 1, figsize=(12, len(groups) * 4), sharex=True)
    colors = plt.cm.tab10(np.linspace(0, 1, len(groups)))

    for i, (group_variability, labels) in enumerate(groups.items()):
        ax = axs[i] if len(groups) > 1 else axs  # Gérer le cas d'un seul groupe
        for label in labels:
            group_cumulative_variability = cumulative_variabilities[label]
            ax.plot(
                range(len(group_cumulative_variability)),
                group_cumulative_variability,
                label=f"{label} (Var: {variabilities[label]:.2f})",
                color=colors[i],
            )

        ax.plot(
            range(len(reference_cumulative_variability)),
            reference_cumulative_variability,
            label="Référence (Cumulative)",
            color='black',
            linestyle='--'
        )

        ax.set_title(f"Groupe {i + 1} (Variabilité ~ {group_variability:.2f})")
        ax.set_xlabel("Temps")
        ax.set_ylabel("Variabilité cumulée")
        ax.legend()
        ax.grid(True)

    plt.tight_layout()
    # Commenter la ligne suivante pour éviter d'afficher le graphique
    #plt.show()

# Fonction principale
def main():
    file_path = 'PPG_data_combined.xlsx'
    sheet_names = ['Sheet1', 'Sheet2']
    columns_per_sheet = [[2, 4, 6, 8, 10, 12, 14, 16], [10, 11, 12, 13]]

    data = load_data(file_path, sheet_names, columns_per_sheet)
    datasets = {
        'Ranked1': data['Sheet1'][0], 'Ranked2': data['Sheet1'][1], 'Ranked3': data['Sheet1'][2],
        'Ranked4': data['Sheet1'][3], 'Ranked5': data['Sheet1'][4], 'Ranked6': data['Sheet1'][5],
        'Ranked7': data['Sheet1'][6], 'Reference': data['Sheet1'][7],
        'Ranked8': data['Sheet2'][0], 'Ranked9': data['Sheet2'][1],
        'Ranked10': data['Sheet2'][2], 'Ranked11': data['Sheet2'][3]
    }

    reference_mean = calculate_reference_mean(datasets['Reference'])
    cumulative_variabilities = {
        label: compute_cumulative_variability(data, reference_mean) for label, data in datasets.items()
    }
    variabilities = calculate_variabilities(datasets, reference_mean)
    tolerance = 0.1 * np.mean(list(variabilities.values()))
    groups = group_datasets_by_variability(variabilities, tolerance)

    ranked_labels = ['Ranked1', 'Ranked2', 'Ranked3', 'Ranked4', 'Ranked5',
                     'Ranked6', 'Ranked7', 'Ranked8', 'Ranked9', 'Ranked10',
                     'Ranked11']

    # Appels de fonctions sans afficher les graphiques
    plot_ranked_series(datasets, reference_mean, ranked_labels)
    reference_cumulative_variability = cumulative_variabilities['Reference']
    plot_grouped_cumulative_variability(groups, cumulative_variabilities, reference_cumulative_variability, variabilities)

    # Afficher les groupes
    for i, (group_variability, labels) in enumerate(groups.items(), 1):
        print(f"Groupe {i} (Variabilité ~ {group_variability:.2f}): {', '.join(labels)}")


# Vérifier si le script est exécuté directement (pas importé)
if __name__ == "__main__":
    main()
